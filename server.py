import urllib.request, urllib.parse, urllib.error
import json, os, uuid, threading, time, io, ssl
from datetime import datetime, timedelta
from flask import Flask, request, send_from_directory, Response

app = Flask(__name__, static_folder='.')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
GAS_URL = 'https://script.google.com/macros/s/AKfycbyQK2h0DgVo00gyaLCg7Aff5F9LinzE22eRjh1zN4IZwqEm_TB0deJ1MGzwUKDP3ZJ5/exec'
API_KEY = 'asdsaAscvlllwiFFa'

def gas_request(params_dict):
    params = urllib.parse.urlencode(params_dict)
    url = f'{GAS_URL}?{params}'
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(url, method='GET')
    req.add_header('User-Agent', 'Mozilla/5.0')
    req.add_header('Accept', 'application/json')
    with urllib.request.urlopen(req, timeout=120, context=ctx) as resp:
        return json.loads(resp.read().decode())

def cleanup_old_selfies():
    try:
        data = gas_request({'key': API_KEY, 'action': 'list', 'sheet': 'FirstLastCalls'})
        entries = data.get('data', [])
        cutoff = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        deleted = 0
        for entry in entries:
            entry_date = entry.get('Date', '')
            if entry_date and entry_date < cutoff:
                gas_request({'key': API_KEY, 'action': 'delete', 'sheet': 'FirstLastCalls', 'idColumn': 'CallID', 'idValue': entry.get('CallID', '')})
                deleted += 1
        if deleted > 0:
            print(f"Cleanup: deleted {deleted} old selfies")
        return deleted
    except Exception as e:
        print(f"Cleanup error: {e}")
        return 0

def scheduled_cleanup():
    while True:
        time.sleep(3600)
        cleanup_old_selfies()

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/<path:path>')
def static_files(path):
    if path.startswith('uploads/'):
        return send_from_directory(UPLOAD_DIR, path.replace('uploads/', ''))
    return send_from_directory(BASE_DIR, path)

@app.route('/upload', methods=['POST'])
def upload_image():
    if 'file' not in request.files:
        return json.dumps({'success': False, 'error': 'No file'}), 400
    f = request.files['file']
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else 'jpg'
    name = str(uuid.uuid4()) + '.' + ext
    f.save(os.path.join(UPLOAD_DIR, name))
    url = request.host_url.rstrip('/') + '/uploads/' + name
    return json.dumps({'success': True, 'url': url})

@app.route('/api/product-template.xlsx')
def api_product_template():
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Name", "Price", "Principal", "ProductID"])
        ws.append(["Sample Product", "99.99", "Principal Name", "PROD001"])
        for i, col in enumerate(['A','B','C','D'], 1):
            ws.column_dimensions[col].width = 20
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'inline; filename="product_template.xlsx"'})
    except ImportError:
        return json.dumps({'success': False, 'error': 'openpyxl not installed on server. Run: pip install openpyxl'}), 500
    except Exception as e:
        return json.dumps({'success': False, 'error': str(e)}), 500

@app.route('/api/import-start', methods=['POST'])
def api_import_start():
    if 'file' not in request.files:
        return json.dumps({'success': False, 'error': 'No file'}), 400
    f = request.files['file']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return json.dumps({'success': False, 'error': 'Empty file'}), 400
        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        for r in ['name', 'price', 'principal', 'productid']:
            if r not in headers: return json.dumps({'success': False, 'error': f'Missing column: {r}'}), 400
        ni = headers.index('name'); pi = headers.index('price'); pri = headers.index('principal'); pidi = headers.index('productid')
        
        # Extract ProductIDs from file
        file_pids = set()
        opts_list = []
        for row in rows[1:]:
            if not row or not row[ni]: continue
            n = str(row[ni]).strip(); p = str(row[pi]).strip() if row[pi] is not None else '0'
            pr = str(row[pri]).strip() if row[pri] else ''; pid = str(row[pidi]).strip() if row[pidi] else ''
            if n and p and pr and pid:
                file_pids.add(pid)
                opts_list.append({'key': API_KEY, 'action': 'create', 'sheet': 'Products', 'Name': n, 'Price': p, 'Principal': pr, 'ProductID': pid})
        
        # Check for duplicates against existing database
        try:
            existing = gas_request({'key': API_KEY, 'action': 'list', 'sheet': 'Products'})
            existing_pids = set()
            for prod in existing.get('data', []):
                pid = (prod.get('ProductID') or '').strip()
                if pid: existing_pids.add(pid)
            
            duplicates = file_pids & existing_pids
            if duplicates:
                dup_list = sorted(duplicates)
                numbered = '\n'.join(f"{i+1}. {d}" for i, d in enumerate(dup_list))
                return json.dumps({'success': False, 'duplicate': True, 
                    'error': f'Duplicate ProductID(s) found:\n{numbered}\n\nPlease remove them from the file and try again.',
                    'duplicates': dup_list}), 400
        except Exception:
            pass  # If check fails, proceed anyway
        
        tid = str(uuid.uuid4())
        import_tasks[tid] = {'opts': opts_list, 'imported': 0, 'errors': [], 'done': False, 'cancelled': False}
        threading.Thread(target=lambda: run_import(tid), daemon=True).start()
        return json.dumps({'success': True, 'task_id': tid, 'total': len(opts_list)})
    except Exception as e:
        return json.dumps({'success': False, 'error': str(e)}), 500

@app.route('/api/customer-template.xlsx')
def api_customer_template():
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Customers"
        ws.append(["CustomerID", "Name", "Address", "Phone", "AgentID"])
        ws.append(["CUST001", "Sample Customer", "123 Street", "09123456789", "AG001"])
        for c in ['A','B','C','D','E']: ws.column_dimensions[c].width = 22
        output = io.BytesIO(); wb.save(output); output.seek(0)
        return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={'Content-Disposition': 'inline; filename="customer_template.xlsx"'})
    except Exception as e:
        return json.dumps({'success': False, 'error': str(e)}), 500

@app.route('/api/import-customers-start', methods=['POST'])
def api_import_customers_start():
    if 'file' not in request.files: return json.dumps({'success': False, 'error': 'No file'}), 400
    f = request.files['file']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(f.read())); ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows: return json.dumps({'success': False, 'error': 'Empty file'}), 400
        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        for r in ['customerid', 'name', 'address', 'phone', 'agentid']:
            if r not in headers: return json.dumps({'success': False, 'error': f'Missing column: {r}'}), 400
        ci = headers.index('customerid'); ni = headers.index('name')
        ai = headers.index('address'); phi = headers.index('phone'); agi = headers.index('agentid')
        
        file_ids = set()
        opts_list = []
        for row in rows[1:]:
            if not row or not row[ni]: continue
            cid = str(row[ci]).strip() if row[ci] else ''
            n = str(row[ni]).strip(); a = str(row[ai]).strip() if row[ai] else ''
            ph = str(row[phi]).strip() if row[phi] else ''; ag = str(row[agi]).strip() if row[agi] else ''
            if cid and n and a and ph and ag:
                file_ids.add(cid)
                opts_list.append({'key': API_KEY, 'action': 'create', 'sheet': 'Customers',
                    'CustomerID': cid, 'Name': n, 'Address': a, 'Phone': ph, 'AgentID': ag})
        
        try:
            existing = gas_request({'key': API_KEY, 'action': 'list', 'sheet': 'Customers'})
            existing_ids = set()
            for c in existing.get('data', []):
                pid = (c.get('CustomerID') or '').strip()
                if pid: existing_ids.add(pid)
            dups = file_ids & existing_ids
            if dups:
                dl = sorted(dups)
                numbered = '\n'.join(f"{i+1}. {d}" for i, d in enumerate(dl))
                return json.dumps({'success': False, 'duplicate': True,
                    'error': f'Duplicate CustomerID(s) found:\n{numbered}\n\nRemove them from the file and try again.',
                    'duplicates': dl}), 400
        except: pass
        
        tid = str(uuid.uuid4())
        import_tasks[tid] = {'opts': opts_list, 'imported': 0, 'errors': [], 'done': False, 'cancelled': False}
        threading.Thread(target=lambda: run_import(tid), daemon=True).start()
        return json.dumps({'success': True, 'task_id': tid, 'total': len(opts_list)})
    except Exception as e:
        return json.dumps({'success': False, 'error': str(e)}), 500

import_tasks = {}
def run_import(tid):
    task = import_tasks.get(tid, {})
    opts_list = task.get('opts', [])
    for i, opts in enumerate(opts_list):
        if task.get('cancelled'): break
        try:
            r = gas_request(opts)
            if r.get('success'): task['imported'] += 1
            else: task['errors'].append(f"Row {i+2}: {r.get('error','fail')}")
        except Exception as e:
            task['errors'].append(f"Row {i+2}: {str(e)[:60]}")
    task['done'] = True

@app.route('/api/import-status/<tid>')
def api_import_status(tid):
    t = import_tasks.get(tid)
    if not t: return json.dumps({'success': False, 'error': 'not found'}), 404
    return json.dumps({'success': True, 'done': t['done'], 'imported': t['imported'], 'total': len(t.get('opts',[])), 'errors': t['errors']})

@app.route('/api/import-cancel/<tid>')
def api_import_cancel(tid):
    t = import_tasks.get(tid)
    if t: t['cancelled'] = True
    return json.dumps({'success': True})

@app.route('/api/process-order', methods=['POST'])
def api_process_order():
    data = request.get_json(force=True) if request.is_json else request.form
    order_id = data.get('orderId', '')
    truck_id = data.get('truckId', '')
    items_json = data.get('itemsJson', '[]')
    
    if not order_id: return json.dumps({'success': False, 'error': 'No orderId'}), 400
    
    # Mark order as Sent
    try:
        gas_request({'key': API_KEY, 'action': 'updateOrderStatus', 'orderId': order_id, 'status': 'Sent'})
    except Exception as e:
        return json.dumps({'success': False, 'error': f'Status update failed: {str(e)}'}), 500
    
    # Deduct truck inventory
    if truck_id:
        try:
            items = json.loads(items_json) if isinstance(items_json, str) else items_json
            for item in items:
                pname = item.get('productName') or item.get('product_name', '')
                qty = item.get('qty') or item.get('quantity', 0)
                if pname and qty:
                    try:
                        gas_request({'key': API_KEY, 'action': 'create', 'sheet': 'TruckInventory',
                            'TruckID': truck_id, 'ProductName': pname, 'Quantity': f'-{int(qty)}'})
                    except: pass
        except: pass
    
    return json.dumps({'success': True, 'message': f'Order {order_id} processed'})

@app.route('/api/cleanup')
def api_cleanup():
    count = cleanup_old_selfies()
    return json.dumps({'success': True, 'deleted': count})

@app.route('/api/', defaults={'path': ''})
@app.route('/api/<path:path>')
def api_proxy(path):
    query = request.query_string.decode('utf-8') if request.query_string else ''
    if request.method == 'POST':
        body = request.get_data(as_text=True)
        if body: query += '&' + body
    url = f'{GAS_URL}?{query}'
    try:
        import ssl
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        gas_req = urllib.request.Request(url, method='GET')
        gas_req.add_header('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        gas_req.add_header('Accept', 'application/json')
        with urllib.request.urlopen(gas_req, timeout=120, context=ctx) as resp:
            return Response(resp.read(), mimetype='application/json')
    except urllib.error.HTTPError as e:
        return json.dumps({'success': False, 'error': f'HTTP {e.code}', 'detail': e.read().decode('utf-8', errors='replace')[:500]}), 502
    except urllib.error.URLError as e:
        return json.dumps({'success': False, 'error': f'Connection error: {str(e.reason)}'}), 502
    except Exception as e:
        return json.dumps({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    import sys
    print("Running initial cleanup...")
    cleanup_old_selfies()
    t = threading.Thread(target=scheduled_cleanup, daemon=True)
    t.start()
    port = int(os.environ.get('PORT', sys.argv[1] if len(sys.argv) > 1 else 8080))
    print(f"Server starting on port {port}")
    app.run(host='0.0.0.0', port=port)
